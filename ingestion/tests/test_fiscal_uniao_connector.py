from __future__ import annotations

import datetime as dt
import io

import openpyxl
import pytest

from ingestion.ckan import CkanResource
from ingestion.connectors.base import ConnectorError
from ingestion.connectors.fiscal_uniao import (
    FiscalUniaoConnector,
    build_default_connector,
    discover_resource,
    parse_to_long_csv,
)


class _FakeResponse:
    def __init__(
        self, status_code: int, content: bytes, payload: dict[str, object] | None = None
    ) -> None:
        self.status_code = status_code
        self.content = content
        self._payload = payload or {}

    def raise_for_status(self) -> None:
        if self.status_code >= 400:
            raise RuntimeError(f"HTTP {self.status_code}")

    def json(self) -> dict[str, object]:
        return self._payload


class _FakeSession:
    def __init__(self, *, package_show: dict[str, object], xlsx_bytes: bytes) -> None:
        self._package_show = package_show
        self._xlsx_bytes = xlsx_bytes

    def get(self, url: str, timeout: float) -> _FakeResponse:
        if "package_show" in url:
            return _FakeResponse(200, b"", self._package_show)
        return _FakeResponse(200, self._xlsx_bytes)


def _package_show(resource_name: str, resource_url: str) -> dict[str, object]:
    return {
        "success": True,
        "result": {
            "resources": [
                {"id": "r1", "name": resource_name, "format": "XLSX", "url": resource_url}
            ]
        },
    }


def _sample_xlsx(*, receita: float, despesa: float, primario: float) -> bytes:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet("1.2")
    sheet.append(["Voltar"])
    sheet.append(["Tabela 1.2. Resultado Primário do Governo Central - Brasil - Mensal"])
    sheet.append(["R$ Milhões - Valores Correntes"])
    sheet.append([])
    sheet.append(["Discriminação", dt.datetime(2026, 6, 1)])
    sheet.append(["1. RECEITA TOTAL 1/", receita + 100])
    sheet.append(["3. RECEITA LÍQUIDA (1-2)", receita])
    sheet.append(["4. DESPESA TOTAL 2/", despesa])
    sheet.append(["5. RESULTADO PRIMÁRIO GOVERNO CENTRAL - ACIMA DA LINHA (3 - 4)", primario])
    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()


_SAMPLE = _sample_xlsx(receita=226305.45, despesa=215522.02, primario=-1234.56)
_RESOURCE_NAME = "Resultado do Tesouro Nacional - Série Histórica - Mensal"
_CKAN_KWARGS = {
    "ckan_base_url": "https://tesourotransparente.gov.br/ckan",
    "ckan_package_id": "ab56485b-9c40-4efb-8563-9ce3e1973c4b",
}


def _session(xlsx_bytes: bytes = _SAMPLE, resource_name: str = _RESOURCE_NAME) -> _FakeSession:
    return _FakeSession(
        package_show=_package_show(resource_name, "https://tesouro/rtn.xlsx"),
        xlsx_bytes=xlsx_bytes,
    )


def _resource(url: str = "https://tesouro/rtn.xlsx") -> CkanResource:
    return CkanResource(
        resource_id="r1", name=_RESOURCE_NAME, format="XLSX", url=url, last_modified=None
    )


def test_discover_resource_resolves_by_name_not_hardcoded_url() -> None:
    resource = discover_resource(_session(), **_CKAN_KWARGS)
    assert resource.url == "https://tesouro/rtn.xlsx"
    assert resource.name == _RESOURCE_NAME


def test_discover_resource_raises_when_name_not_found() -> None:
    with pytest.raises(ConnectorError):
        discover_resource(_session(resource_name="Some Other Resource"), **_CKAN_KWARGS)


def test_build_default_connector_wires_discovery_and_connector() -> None:
    connector = build_default_connector(_session(), **_CKAN_KWARGS)
    ref = connector.discover()
    assert ref.resource_url == "https://tesouro/rtn.xlsx"
    assert ref.resource_format == "xlsx"


def test_download_writes_original_xlsx_bytes_unmodified(tmp_path) -> None:
    connector = FiscalUniaoConnector(session=_session(), resource=_resource())
    dest = str(tmp_path / "resource.xlsx")
    result = connector.download(connector.discover(), dest)
    assert tmp_path.joinpath("resource.xlsx").read_bytes() == _SAMPLE
    assert result.bytes_downloaded == len(_SAMPLE)


def test_download_supports_file_scheme_for_tests_without_network(tmp_path) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(_SAMPLE)
    connector = FiscalUniaoConnector(session=None, resource=_resource(f"file://{source}"))
    dest = str(tmp_path / "resource.xlsx")
    result = connector.download(connector.discover(), dest)
    assert result.bytes_downloaded == len(_SAMPLE)
    assert result.http_status == 200


def test_validate_accepts_real_shaped_file(tmp_path) -> None:
    dest = tmp_path / "resource.xlsx"
    dest.write_bytes(_SAMPLE)
    connector = FiscalUniaoConnector(session=_session(), resource=_resource())
    connector.validate(str(dest))


def test_validate_rejects_file_missing_target_sheet(tmp_path) -> None:
    workbook = openpyxl.Workbook()
    buf = io.BytesIO()
    workbook.save(buf)
    dest = tmp_path / "resource.xlsx"
    dest.write_bytes(buf.getvalue())
    connector = FiscalUniaoConnector(session=_session(), resource=_resource())
    with pytest.raises(ConnectorError):
        connector.validate(str(dest))


def test_parse_to_long_csv_pivots_wide_table_to_long_rows() -> None:
    csv_bytes = parse_to_long_csv(_SAMPLE)
    text = csv_bytes.decode("utf-8")
    lines = text.strip().splitlines()
    assert lines[0] == "metric_id,reference_period,value_millions"
    rows = dict((line.split(",")[0], line.split(",")[2]) for line in lines[1:])
    assert rows["fiscal_receita"] == "226305.45"
    assert rows["fiscal_despesa"] == "215522.02"
    assert rows["fiscal_primario"] == "-1234.56"


def test_parse_to_long_csv_raises_when_row_label_missing() -> None:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet("1.2")
    for _ in range(4):
        sheet.append([])
    sheet.append(["Discriminação", dt.datetime(2026, 6, 1)])
    sheet.append(["1. RECEITA TOTAL 1/", 100])
    buf = io.BytesIO()
    workbook.save(buf)
    with pytest.raises(ConnectorError):
        parse_to_long_csv(buf.getvalue())


def test_checkpoint_always_proceeds_when_no_prior_hash_recorded() -> None:
    connector = FiscalUniaoConnector(session=_session(), resource=_resource())
    ref = connector.discover()
    assert connector.checkpoint(ref, "any-hash") is True
