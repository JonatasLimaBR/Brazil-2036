from __future__ import annotations

import csv
import datetime as dt
import hashlib
import io
from collections.abc import Sequence
from decimal import Decimal
from pathlib import Path
from typing import Protocol

import openpyxl

from ingestion.ckan import CkanResource, list_resources
from ingestion.connectors.base import (
    ConnectorError,
    DownloadResult,
    ResourceRef,
    retry_with_backoff,
)

# "Resultado do Tesouro Nacional - Série Histórica" (Tesouro Transparente/CKAN,
# package id ab56485b-9c40-4efb-8563-9ce3e1973c4b, org CESEF, license ODbL).
# Confirmed real by downloading and opening the file (DESIGN §0): 27 sheets,
# sheet "1.2" is "Resultado Primário do Governo Central - Brasil - Mensal", a
# wide table (1 row per rubric x 1 column per month, R$ millions, current
# values, Jan/1997-Jul/2026 in the version inspected). The download URL's
# filename changes every month (e.g. "seriehistoricajul26.xlsx") but the CKAN
# resource is resolved by name via package_show, never by a hardcoded filename
# pattern -- same principle as the INSS agency-portal discovery.
_RESOURCE_NAME = "Resultado do Tesouro Nacional - Série Histórica - Mensal"
_TARGET_SHEET = "1.2"
_HEADER_ROW_INDEX = 4

# metric_id -> row label prefix. Receita maps to "RECEITA LÍQUIDA" (post
# revenue-sharing transfers), not "RECEITA TOTAL": the source's own published
# primary result is Receita Líquida minus Despesa Total, and using Receita
# Total here would make our fiscal_receita - fiscal_despesa disagree with the
# fiscal_primario the source itself publishes (DESIGN D4).
_TARGET_ROWS: tuple[tuple[str, str], ...] = (
    ("fiscal_receita", "3. RECEITA LÍQUIDA"),
    ("fiscal_despesa", "4. DESPESA TOTAL"),
    ("fiscal_primario", "5. RESULTADO PRIMÁRIO GOVERNO CENTRAL - ACIMA DA LINHA"),
)

_CSV_HEADER = ("metric_id", "reference_period", "value_millions")
_FILE_SCHEME = "file://"


class HttpResponse(Protocol):
    status_code: int
    content: bytes

    def raise_for_status(self) -> None: ...

    def json(self) -> dict[str, object]: ...


class HttpSession(Protocol):
    def get(self, url: str, timeout: float) -> HttpResponse: ...


def discover_resource(
    session: HttpSession, *, ckan_base_url: str, ckan_package_id: str, timeout: float = 30.0
) -> CkanResource:
    """Resolve the current CkanResource for the RTN monthly series by name.

    Never by a hardcoded filename pattern: the download URL's filename changes
    every month (e.g. "seriehistoricajul26.xlsx") but the resource is looked
    up fresh via package_show each time -- same principle as the INSS
    agency-portal discovery.
    """
    resources = list_resources(
        session, base_url=ckan_base_url, package_id=ckan_package_id, timeout=timeout
    )
    matches = [r for r in resources if r.name == _RESOURCE_NAME]
    if not matches:
        names = [r.name for r in resources]
        raise ConnectorError(
            f"resource {_RESOURCE_NAME!r} not found in package {ckan_package_id!r}; "
            f"available: {names!r}"
        )
    return matches[0]


class FiscalUniaoConnector:
    """Central Government revenue/expense/primary result, wide monthly series.

    Takes an already-resolved CkanResource (via `discover_resource`), same
    separation of concerns as `InssEmitidosConnector`: this class only knows
    how to fetch/validate/parse one given resource, not how to find it.

    RAW stores the original XLSX bytes unmodified (DESIGN D6 -- unlike the
    INSS Indeferidos connector, which stores only the converted CSV). The
    wide-to-long pivot happens separately, in `parse_to_long_csv`, called by
    the pipeline after RAW has already captured the untouched source file.
    """

    dataset_id = "fiscal_uniao"

    def __init__(
        self,
        *,
        session: HttpSession | None,
        resource: CkanResource,
        max_retries: int = 4,
        backoff_seconds: float = 2.0,
        request_timeout: float = 120.0,
    ) -> None:
        self._session = session
        self._resource = resource
        self._max_retries = max_retries
        self._backoff_seconds = backoff_seconds
        self._request_timeout = request_timeout

    def discover(self) -> ResourceRef:
        return ResourceRef(
            dataset_id=self.dataset_id,
            resource_url=self._resource.url,
            resource_format="xlsx",
            resource_hash=None,
        )

    def metadata(self, ref: ResourceRef) -> dict[str, str]:
        return {"resource_id": self._resource.resource_id, "resource_name": self._resource.name}

    def download(self, ref: ResourceRef, dest: str) -> DownloadResult:
        if ref.resource_url.startswith(_FILE_SCHEME):
            data = Path(ref.resource_url[len(_FILE_SCHEME) :]).read_bytes()
            Path(dest).write_bytes(data)
            return DownloadResult(
                local_path=dest,
                content_sha256=hashlib.sha256(data).hexdigest(),
                http_status=200,
                bytes_downloaded=len(data),
                attempts=1,
                attempt_errors=[],
            )

        errors: list[str] = []

        def _fetch() -> HttpResponse:
            assert self._session is not None
            response = self._session.get(ref.resource_url, timeout=self._request_timeout)
            response.raise_for_status()
            return response

        response = retry_with_backoff(
            _fetch,
            max_attempts=self._max_retries,
            backoff_seconds=self._backoff_seconds,
            errors=errors,
        )
        data = response.content
        with open(dest, "wb") as handle:
            handle.write(data)
        return DownloadResult(
            local_path=dest,
            content_sha256=hashlib.sha256(data).hexdigest(),
            http_status=response.status_code,
            bytes_downloaded=len(data),
            attempts=len(errors) + 1,
            attempt_errors=errors,
        )

    def validate(self, local_path: str) -> None:
        with open(local_path, "rb") as handle:
            _extract_series(handle.read())

    def checkpoint(self, ref: ResourceRef, content_sha256: str) -> bool:
        return ref.resource_hash != content_sha256


def build_default_connector(
    session: HttpSession, *, ckan_base_url: str, ckan_package_id: str
) -> FiscalUniaoConnector:
    resource = discover_resource(
        session, ckan_base_url=ckan_base_url, ckan_package_id=ckan_package_id
    )
    return FiscalUniaoConnector(session=session, resource=resource)


def parse_to_long_csv(xlsx_bytes: bytes) -> bytes:
    """Pivot the wide monthly table into a long CSV Bronze can LOAD DATA from.

    One row per (metric_id, reference_period); value stays in R$ millions as
    published -- the millions-to-reais conversion (DESIGN D7) happens in the
    Silver SQL, not here.
    """
    series = _extract_series(xlsx_bytes)
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(_CSV_HEADER)
    for metric_id, period, value in series:
        writer.writerow((metric_id, period.isoformat(), str(value)))
    return buffer.getvalue().encode("utf-8")


def _extract_series(xlsx_bytes: bytes) -> list[tuple[str, dt.date, Decimal]]:
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        if _TARGET_SHEET not in workbook.sheetnames:
            raise ConnectorError(
                f"sheet {_TARGET_SHEET!r} not found; available: {workbook.sheetnames!r}"
            )
        sheet = workbook[_TARGET_SHEET]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) <= _HEADER_ROW_INDEX:
            raise ConnectorError(
                f"sheet {_TARGET_SHEET!r} has fewer than {_HEADER_ROW_INDEX + 1} rows"
            )

        header = rows[_HEADER_ROW_INDEX]
        periods: list[tuple[int, dt.date]] = []
        for col_idx, cell in enumerate(header):
            if col_idx == 0:
                continue
            if isinstance(cell, dt.datetime):
                periods.append((col_idx, cell.date().replace(day=1)))
            elif isinstance(cell, dt.date):
                periods.append((col_idx, cell.replace(day=1)))

        if not periods:
            raise ConnectorError(f"no month columns found in header row of sheet {_TARGET_SHEET!r}")

        series: list[tuple[str, dt.date, Decimal]] = []
        for metric_id, label_prefix in _TARGET_ROWS:
            row = _find_row(rows, label_prefix)
            for col_idx, period in periods:
                value_cell = row[col_idx] if col_idx < len(row) else None
                if value_cell is None or value_cell == "":
                    continue
                series.append((metric_id, period, Decimal(str(value_cell))))
        return series
    finally:
        workbook.close()


def _find_row(rows: Sequence[tuple[object, ...]], label_prefix: str) -> tuple[object, ...]:
    for row in rows:
        label = row[0] if row else None
        if isinstance(label, str) and label.strip().startswith(label_prefix):
            return row
    raise ConnectorError(
        f"row starting with {label_prefix!r} not found in sheet {_TARGET_SHEET!r} "
        "-- source layout changed"
    )
